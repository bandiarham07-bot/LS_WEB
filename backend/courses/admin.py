from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from .models import (
    Assignment,
    AssignmentSubmission,
    User,
    CourseStatus,
    ContentPage,
    ContentBlock,
    UserProgress,
)


def format_datetime(value):
    if not value:
        return ''
    return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S')


def excel_safe_sheet_title(title, existing_titles):
    clean_title = ''.join(
        character if character not in r'[]:*?/\\' else '-'
        for character in title
    ).strip() or 'Assignment'
    clean_title = clean_title[:31]

    candidate = clean_title
    counter = 2
    while candidate in existing_titles:
        suffix = f' {counter}'
        candidate = f'{clean_title[:31 - len(suffix)]}{suffix}'
        counter += 1
    return candidate


def build_assignment_submissions_workbook(assignments):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    existing_titles = set()

    for assignment in assignments:
        sheet_title = excel_safe_sheet_title(assignment.page.title, existing_titles)
        existing_titles.add(sheet_title)
        worksheet = workbook.create_sheet(title=sheet_title)
        worksheet.append([
            'Assignment',
            'Roll number',
            'Student email',
            'Student name',
            'GitHub repository',
            'GitHub repository URL',
            'Grade awarded',
            'Grade total',
            'Evaluator note',
            'Graded at',
            'Submitted at',
            'Updated at',
        ])

        submissions = (
            assignment.submissions
            .select_related('user')
            .order_by('roll_number', 'user__email')
        )
        for submission in submissions:
            student_name = ' '.join(
                part for part in [
                    submission.user.first_name,
                    submission.user.last_name,
                ] if part
            )
            worksheet.append([
                assignment.page.title,
                submission.roll_number,
                submission.user.email,
                student_name,
                submission.github_repo_name,
                submission.github_repo_url,
                submission.grade_awarded,
                submission.grade_total,
                submission.evaluator_note,
                format_datetime(submission.graded_at),
                format_datetime(submission.submitted_at),
                format_datetime(submission.updated_at),
            ])

        for column_cells in worksheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 12),
                50,
            )

    if not workbook.sheetnames:
        workbook.create_sheet(title='Submissions')

    return workbook


@admin.register(User)
class CustomUserAdmin(UserAdmin):
    list_display = ['email', 'first_name', 'last_name', 'is_staff', 'date_joined']
    ordering = ['email']


class ContentBlockInline(admin.StackedInline):
    """
    StackedInline puts every field on its own row so nothing gets hidden.
    Fields are grouped logically: order/type first, then the three content
    fields (title, body, url) each labelled clearly.
    """
    model = ContentBlock
    extra = 1
    fields = ['order', 'type', 'title', 'body', 'url']
    ordering = ['order']

    # Clearer help text so it's obvious which field to use per type
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        form.base_fields['body'].help_text = 'Used by Text blocks — write your content here.'
        form.base_fields['url'].help_text  = (
            'Used by YouTube and Document blocks. '
            'For YouTube paste the full watch URL '
            '(e.g. https://www.youtube.com/watch?v=...). '
            'For documents paste a direct file or Google Drive link.'
        )
        form.base_fields['order'].help_text = (
            'Controls the display order within this page. '
            'Lower numbers appear first. Block 1 is the first thing students see.'
        )
        return form


class AssignmentInline(admin.StackedInline):
    model = Assignment
    extra = 0
    max_num = 1
    fields = ['details', 'deliverables', 'due_date', 'opens_at', 'weightage']


@admin.register(ContentPage)
class ContentPageAdmin(admin.ModelAdmin):
    """
    • list_editable on 'order' lets you change page order directly from
      the list view without opening each page individually.
    • list_display_links must exclude 'order' when it is list_editable.
    • ordering + list_filter make it easy to find and reorder pages.
    """
    list_display         = ['section', 'order', 'title', 'next_page']
    list_display_links   = ['title']          # 'order' must NOT be a link when editable
    list_editable        = ['order']          # ← edit order numbers inline on list page
    list_filter          = ['section']
    ordering             = ['section', 'order']
    inlines              = [AssignmentInline, ContentBlockInline]

    # Show next_page choices filtered to same section for clarity
    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        if 'order' in form.base_fields:
            form.base_fields['order'].help_text = (
                'Page position within its section. '
                '1 = first page students land on in this section.'
            )
        return form


@admin.register(ContentBlock)
class ContentBlockAdmin(admin.ModelAdmin):
    """
    Standalone view so you can also edit blocks directly if needed.
    """
    list_display  = ['page', 'order', 'type', 'title']
    list_filter   = ['type', 'page__section']
    ordering      = ['page__section', 'page__order', 'order']
    list_editable = ['order']
    list_display_links = ['title']


@admin.register(Assignment)
class AssignmentAdmin(admin.ModelAdmin):
    list_display = ['page', 'due_date', 'opens_at', 'weightage', 'is_open']
    list_filter = ['due_date', 'opens_at']
    search_fields = ['page__title', 'details', 'deliverables']
    readonly_fields = ['is_open']
    actions = ['export_submissions_excel']

    @admin.action(description='Download submission data as Excel')
    def export_submissions_excel(self, request, queryset):
        assignments = queryset.select_related('page').prefetch_related('submissions__user')
        workbook = build_assignment_submissions_workbook(assignments)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            'attachment; filename="assignment-submissions.xlsx"'
        )
        workbook.save(response)
        return response


@admin.register(AssignmentSubmission)
class AssignmentSubmissionAdmin(admin.ModelAdmin):
    list_display = [
        'assignment',
        'user',
        'roll_number',
        'github_repo_name',
        'grade_display',
        'graded_at',
        'github_repo_url',
        'submitted_at',
        'updated_at',
    ]
    list_filter = ['assignment', 'submitted_at', 'graded_at']
    search_fields = [
        'assignment__page__title',
        'user__email',
        'roll_number',
        'github_repo_name',
        'github_repo_url',
        'evaluator_note',
    ]
    readonly_fields = ['submitted_at', 'updated_at']
    fields = [
        'assignment',
        'user',
        'roll_number',
        'github_repo_url',
        'github_repo_name',
        'grade_awarded',
        'grade_total',
        'evaluator_note',
        'graded_at',
        'submitted_at',
        'updated_at',
    ]

    @admin.display(description='Grade')
    def grade_display(self, obj):
        if obj.grade_awarded is None:
            return 'Not graded'
        if obj.grade_total is None:
            return obj.grade_awarded
        return f'{obj.grade_awarded}/{obj.grade_total}'

    def save_model(self, request, obj, form, change):
        grade_fields = {'grade_awarded', 'grade_total', 'evaluator_note'}
        if grade_fields.intersection(form.changed_data) and not obj.graded_at:
            obj.graded_at = timezone.now()
        super().save_model(request, obj, form, change)


@admin.register(CourseStatus)
class CourseStatusAdmin(admin.ModelAdmin):
    list_display = ['current_week', 'current_topic', 'updated_at']


@admin.register(UserProgress)
class UserProgressAdmin(admin.ModelAdmin):
    list_display   = ['user', 'last_page', 'updated_at']
    readonly_fields = ['updated_at']
